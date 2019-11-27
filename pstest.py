from bs4 import BeautifulSoup
import csv
import openpyxl
from openpyxl import load_workbook
import requests

seller = input("Input seller's name: ")


def get_html(url):
	
	r = requests.get(url)
	return r.text

def page_check(html):
	soup = BeautifulSoup(html, 'lxml')
	try:
		next_page_check = soup.find('td', class_ = 'pagn-next').find('a').get('class')[-1]
	except: 
		next_page_check = None
	return next_page_check

def get_links(html):
	soup = BeautifulSoup(html, 'lxml')
	urls = soup.find('div', id="CenterPanel").find_all('li', class_ = 'sresult')
	links = []

	for link in urls:
		a = link.find('div', class_ = 'lvpicinner full-width picW').find('a').get('href')
		html = get_html(a)
		get_page_data(html)
	return links 
	


def get_page_data(html):
	soup = BeautifulSoup(html, 'lxml')
	pricef = soup.find('div', id="mainContent")
	try:
		title = soup.find('div', id = 'LeftSummaryPanel').find('h1', class_ = 'it-ttl').text.strip().split('   ')[-1]
		print(title)
	except:
		title = ''
	try:
		url = soup.find('link', hreflang = 'en-us').get('href')
		print(url)
	except:
		url = ''
	try:
		price = pricef.find('span', itemprop="price").text.split(' ')
		ship_price = pricef.find('span', id="fshippingCost").find('span').text

		if price[0] != 'US':
			price = pricef.find('span', id = "convbinPrice").text.split('$')[1].split('(')[0]
			if ship_price != 'FREE':
				ship_price = pricef.find('span', id="convetedPriceId").text.split('$')[1]	
			else:
				ship_price = 0

			total_price = float(price) + float(ship_price)
			print(price , ' + ', ship_price, ' = ', total_price)

		else:
			price = price[1].split('$')
			if ship_price != 'FREE':
				ship_price = ship_price.split('$')[1]
			else:	
				ship_price = 0
			total_price = float(price[1]) + float(ship_price)
			print(total_price)
	
	except:
		total_price = ''
	try:
		image_url = soup.find('img', itemprop="image").get('src')
		print(image_url)
	except:
		image_url = ''
	try:
		sold = soup.find('a', class_="vi-txt-underline").text.split(' ')[0]
		print(sold + ' units has been solded')
	except:
		sold = ''
	try:
		available = soup.find('span', id="qtySubTxt").text.split()[0]
		print(available + ' units are stil available')
	except:
		available = ''

	data = {'title': title,
			'url': url,
			'price': str(total_price),
			'image_url': image_url,
			'sold': sold,
			'available': available}

	excel_export(data)

	return data

def excel_export(data):

	wb = load_workbook(seller + '.xlsx')

	title = data['title']
	price = data['price']
	url = data['url']
	image_url = data['image_url']
	sold = data['sold']
	available = data['available']

	ws = wb.active

	last_row = ws.max_row + 1
	print('№'+ str(ws.max_row))

	ws.cell(row = last_row, column = 1, value = title)
	ws.cell(row = last_row, column = 2, value = price)
	ws.cell(row = last_row, column = 3).hyperlink = url
	ws.cell(row = last_row, column = 3).value = url
	ws.cell(row = last_row, column = 3).style = "Hyperlink"
	ws.cell(row = last_row, column = 4).hyperlink = image_url
	ws.cell(row = last_row, column = 4).value = image_url
	ws.cell(row = last_row, column = 4).style = "Hyperlink"
	ws.cell(row = last_row, column = 5, value = sold)
	ws.cell(row = last_row, column = 6, value = available)

	wb.save(seller + '.xlsx')	

def main():

	page = 1
	page_checkk = 'next'
	rows = -1
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.column_dimensions["A"].width = 60
	ws.column_dimensions["B"].width = 30
	ws['A1'] = 'Title'
	ws['B1'] = 'Price(inc. shipping cost)'
	ws['C1'] = 'URL'
	ws['D1'] = 'Image URL'
	ws['E1'] = 'Sold'
	ws['F1'] = 'Available'

	try:
		wb.save(seller + '.xlsx')
	except:
		print('Close your current excel file')

	while page_checkk == 'next':
		url_gen = 'https://www.ebay.com/sch/m.html?_nkw=&_armrs=1&_from=&_ssn='+ seller +'&_pgn=' + str(page)
		html = get_html(url_gen)
		data = get_page_data(html)
		get_links(html)
		page_checkk = page_check(html)
		if page_checkk == None: page_checkk = ''
		print(page_checkk)
		page += 1

if __name__ == '__main__':
  	main()